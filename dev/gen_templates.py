from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

OUT = r'C:\Users\Administrator\Documents\Codex\2026-08-24\build-x20\outputs\teacher-workbench\assets\templates'
os.makedirs(OUT, exist_ok=True)


def style_header(ws, ncols):
    fill = PatternFill('solid', fgColor='E6F2EC')
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='3F7D66')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def add_notes(wb, title, lines):
    notes = wb.create_sheet(title)
    for line in lines:
        notes.append([line])
    set_widths(notes, [72])


# ---------- 花名册模板 ----------
wb = Workbook()
ws = wb.active
ws.title = '花名册'
ws.append(['姓名', '性别', '小组', '排', '列', '班委职务', '家长姓名', '联系电话'])
ws.append(['张小明', '男', '1', '1', '1', '班长', '张某某', '13800000000'])
style_header(ws, 8)
set_widths(ws, [12, 8, 8, 8, 8, 14, 12, 16])
add_notes(wb, '填写说明', [
    '花名册填写说明：',
    '1. 第一行为表头，请勿删除或修改。',
    '2. 「排」「列」为教室座位（1-6 排 × 1-8 列），可留空表示未安排座位。',
    '3. 「小组」填写 1-8。',
    '4. 填好后保存文件，在应用中进入「导入数据 → 花名册」选择本文件即可。',
    '5. 也支持 CSV 格式，列顺序同上。'
])
wb.save(os.path.join(OUT, '花名册模板.xlsx'))

# ---------- 课程表模板 ----------
wb = Workbook()
ws = wb.active
ws.title = '课程表'
ws.append(['时段', '周一', '周二', '周三', '周四', '周五'])
for p in ['早读', '第1节', '第2节', '第3节', '第4节', '第5节', '第6节', '第7节', '第8节']:
    ws.append([p, '', '', '', '', ''])
style_header(ws, 6)
set_widths(ws, [12, 12, 12, 12, 12, 12])
add_notes(wb, '填写说明', [
    '课程表填写说明：',
    '1. 首行为日期表头（周一~周五）。',
    '2. 第一列填写时段：早读、第1节…第8节，可增删行。',
    '3. 课程格填写「科目」或「科目/教师」，例如：数学/王老师。',
    '4. 留空的格子表示空课；含「午休」的行会被自动跳过。',
    '5. 填好后保存文件，在应用中进入「导入数据 → 课程表」选择本文件即可。'
])
wb.save(os.path.join(OUT, '课程表模板.xlsx'))

print('templates:', sorted(os.listdir(OUT)))
