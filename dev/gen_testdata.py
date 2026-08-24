# -*- coding: utf-8 -*-
"""生成测试用花名册与课程表假数据（CSV + Excel）。"""
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = r'C:\Users\Administrator\Documents\Codex\2026-08-24\build-x20\outputs\teacher-workbench\test-data'
os.makedirs(OUT, exist_ok=True)

# ---------------- 花名册假数据（26 人：24 人已安排座位 + 2 人未安排） ----------------
ROSTER_HEADER = ['姓名', '性别', '小组', '排', '列', '班委职务', '家长姓名', '联系电话']
ROSTER_ROWS = [
    ['陈浩宇', '男', '1', '1', '1', '班长', '陈国强', '13912340001'],
    ['刘思琪', '女', '1', '1', '2', '学习委员', '刘明', '13912340002'],
    ['王雨欣', '女', '1', '2', '1', '语文课代表', '王海', '13912340003'],
    ['李俊豪', '男', '1', '2', '2', '数学课代表', '李伟', '13912340004'],
    ['张雅婷', '女', '1', '3', '1', '英语课代表', '张军', '13912340005'],
    ['赵子轩', '男', '1', '3', '2', '', '赵磊', '13912340006'],
    ['孙梦洁', '女', '2', '1', '3', '纪律委员', '孙涛', '13912340007'],
    ['周宇航', '男', '2', '1', '4', '体育委员', '周平', '13912340008'],
    ['吴佳怡', '女', '2', '2', '3', '', '吴斌', '13912340009'],
    ['郑浩然', '男', '2', '2', '4', '', '郑军', '13912340010'],
    ['冯可欣', '女', '2', '3', '3', '宣传委员', '冯强', '13912340011'],
    ['蒋天佑', '男', '2', '3', '4', '', '蒋华', '13912340012'],
    ['林语桐', '女', '3', '4', '5', '文艺委员', '林峰', '13912340013'],
    ['黄泽宇', '男', '3', '4', '6', '', '黄刚', '13912340014'],
    ['徐欣妍', '女', '3', '5', '5', '', '徐斌', '13912340015'],
    ['何俊杰', '男', '3', '5', '6', '劳动委员', '何军', '13912340016'],
    ['马静怡', '女', '3', '6', '5', '', '马涛', '13912340017'],
    ['谢文博', '男', '3', '6', '6', '', '谢平', '13912340018'],
    ['罗婉清', '女', '4', '4', '7', '生活委员', '罗海', '13912340019'],
    ['曹思远', '男', '4', '4', '8', '', '曹军', '13912340020'],
    ['唐欣悦', '女', '4', '5', '7', '', '唐伟', '13912340021'],
    ['袁志强', '男', '4', '5', '8', '', '袁华', '13912340022'],
    ['高梦琪', '女', '4', '6', '7', '', '高峰', '13912340023'],
    ['潘俊杰', '男', '4', '6', '8', '', '潘军', '13912340024'],
    ['郭晓彤', '女', '1', '', '', '', '郭磊', '13912340025'],
    ['邓凯文', '男', '2', '', '', '', '邓涛', '13912340026'],
]

# ---------------- 课程表假数据（周一~周五 × 早读+8节，含午休行与空课） ----------------
SCHEDULE_HEADER = ['时段', '周一', '周二', '周三', '周四', '周五']
SCHEDULE_ROWS = [
    ['早读', '语文/李老师', '英语/张老师', '语文/李老师', '英语/张老师', '语文/李老师'],
    ['第1节', '数学/王老师', '数学/王老师', '英语/张老师', '物理/刘老师', '数学/王老师'],
    ['第2节', '语文/李老师', '英语/张老师', '数学/王老师', '语文/李老师', '英语/张老师'],
    ['第3节', '英语/张老师', '物理/刘老师', '语文/李老师', '数学/王老师', '化学/陈老师'],
    ['第4节', '物理/刘老师', '化学/陈老师', '体育/孙老师', '英语/张老师', '语文/李老师'],
    ['午休', '', '', '', '', ''],
    ['第5节', '化学/陈老师', '语文/李老师', '物理/刘老师', '化学/陈老师', '体育'],
    ['第6节', '数学/王老师', '体育/孙老师', '美术/冯老师', '体育/孙老师', '物理/刘老师'],
    ['第7节', '道德与法治/赵老师', '历史/周老师', '道德与法治/赵老师', '历史/周老师', '班会/张老师'],
    ['第8节', '信息技术/吴老师', '音乐/郑老师', '自习', '班会/张老师', ''],
]


def style_header(ws, ncols):
    fill = PatternFill('solid', fgColor='E6F2EC')
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='3F7D66')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_csv(path, header, rows):
    with open(path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def write_xlsx(path, sheet_name, header, rows, widths):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(header)
    for r in rows:
        ws.append(r)
    style_header(ws, len(header))
    set_widths(ws, widths)
    wb.save(path)


write_csv(os.path.join(OUT, '花名册-测试数据.csv'), ROSTER_HEADER, ROSTER_ROWS)
write_csv(os.path.join(OUT, '课程表-测试数据.csv'), SCHEDULE_HEADER, SCHEDULE_ROWS)
write_xlsx(os.path.join(OUT, '花名册-测试数据.xlsx'), '花名册', ROSTER_HEADER, ROSTER_ROWS, [12, 8, 8, 8, 8, 14, 12, 16])
write_xlsx(os.path.join(OUT, '课程表-测试数据.xlsx'), '课程表', SCHEDULE_HEADER, SCHEDULE_ROWS, [14, 16, 16, 16, 16, 16])

print('generated:')
for name in sorted(os.listdir(OUT)):
    print(' -', name, os.path.getsize(os.path.join(OUT, name)), 'bytes')

# 自检：重新打开 Excel 确认可读
from openpyxl import load_workbook
for f in ['花名册-测试数据.xlsx', '课程表-测试数据.xlsx']:
    wb = load_workbook(os.path.join(OUT, f))
    ws = wb.active
    print('verify', f, '->', ws.title, ws.max_row, 'rows x', ws.max_column, 'cols')
