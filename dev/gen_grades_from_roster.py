# -*- coding: utf-8 -*-
"""根据一份花名册（应用导入格式）生成配套的成绩测试数据。
用法：python dev/gen_grades_from_roster.py <花名册.xlsx|csv> [输出文件名前缀]
"""
import csv
import os
import random
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, '..', 'test-data')
os.makedirs(OUT, exist_ok=True)

SUBJECTS = ['语文', '数学', '英语', '物理', '化学', '生物', '道德与法治', '历史', '地理']


def read_roster(path):
    if path.lower().endswith('.csv'):
        with open(path, encoding='utf-8-sig') as f:
            rows = list(csv.reader(f))
    else:
        ws = load_workbook(path, data_only=True, read_only=True).active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    data = rows[1:] if rows and str(rows[0][0] or '').strip() in ('姓名', '学籍号') else rows
    names = [str(r[0]).strip() for r in data if r and str(r[0]).strip()]
    return names


def score_for(name, subj_idx):
    rnd = random.Random(sum(ord(c) for c in name) * 31 + subj_idx * 17 + 7)
    base = rnd.randint(55, 90)
    return max(40, min(100, base + rnd.randint(-12, 12)))


def build_rows(names, blank=None):
    rows = []
    for name in names:
        row = [name]
        for j in range(len(SUBJECTS)):
            if blank and blank(name, SUBJECTS[j]):
                row.append('')
            else:
                row.append(score_for(name, j))
        rows.append(row)
    return rows


def style_header(ws, ncols):
    fill = PatternFill('solid', fgColor='E6F2EC')
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='3F7D66')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    if len(sys.argv) < 2:
        print('用法: python dev/gen_grades_from_roster.py <花名册.xlsx|csv> [输出前缀]')
        sys.exit(1)
    path = sys.argv[1]
    prefix = sys.argv[2] if len(sys.argv) > 2 else '成绩-测试数据'
    names = read_roster(path)
    header = ['姓名'] + SUBJECTS
    rows = build_rows(names, blank=lambda n, s: n == names[-1] and s == '物理')

    csv_path = os.path.join(OUT, f'{prefix}.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)

    xlsx_path = os.path.join(OUT, f'{prefix}.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = '成绩'
    ws.append(header)
    for r in rows:
        ws.append(r)
    style_header(ws, len(header))
    set_widths(ws, [14] + [9] * len(SUBJECTS))
    notes = wb.create_sheet('填写说明')
    for line in [
        '成绩测试数据说明：',
        f'1. 按花名册生成，共 {len(names)} 名学生，姓名完全一致。',
        '2. 在应用中进入「导入数据 → 成绩」选择本文件，将自动新建一场考试。',
        f'3. 「{names[-1]}-物理」留空，用于测试缺考单元格的处理。'
    ]:
        notes.append([line])
    set_widths(notes, [72])
    wb.save(xlsx_path)
    print(f'generated {len(names)} students -> {os.path.basename(csv_path)} / {os.path.basename(xlsx_path)}')


if __name__ == '__main__':
    main()
