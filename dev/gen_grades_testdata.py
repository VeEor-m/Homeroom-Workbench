# -*- coding: utf-8 -*-
"""生成成绩分析测试数据（CSV + Excel）。
包含三个版本：
  1. 成绩-测试数据           —— 姓名列，对应 js/data.js 示例花名册（48 人）
  2. 成绩-测试数据-学籍号版   —— 首列学籍号，姓名被改动时也能匹配
  3. 成绩-测试数据-花名册测试版 —— 对应 花名册-测试数据.csv（26 人）
"""
import csv
import os
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, '..', 'test-data')
os.makedirs(OUT, exist_ok=True)

SUBJECTS = ['语文', '数学', '英语', '物理', '化学', '生物', '道德与法治', '历史', '地理']

# 从 js/data.js 提取示例花名册姓名（保证导入时能匹配上）
src = open(os.path.join(BASE, '..', 'js', 'data.js'), encoding='utf-8').read()
block_start = src.index('const STUDENTS = [')
block_end = src.index('];', block_start)
names = re.findall(r"name:\s*'([^']+)'", src[block_start:block_end])

# 花名册测试数据（26 人）的姓名，来自 test-data/花名册-测试数据.csv
roster_csv = os.path.join(BASE, '..', 'test-data', '花名册-测试数据.csv')
roster_names = []
if os.path.exists(roster_csv):
    with open(roster_csv, encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if row and row[0].strip():
                roster_names.append(row[0].strip())


def score_for(name, subj_idx):
    rnd = random.Random(sum(ord(c) for c in name) * 31 + subj_idx * 17 + 7)
    base = rnd.randint(55, 90)
    return max(40, min(100, base + rnd.randint(-12, 12)))


def build_rows(name_list, blank=None):
    rows = []
    for name in name_list:
        row = [name]
        for j in range(len(SUBJECTS)):
            if blank and blank(name, SUBJECTS[j]):
                row.append('')
            else:
                row.append(score_for(name, j))
        rows.append(row)
    return rows


def style_header(ws, ncols):
    fill = PatternFill('solid', fgColor='E6F2EC')
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='3F7D66')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')


def set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_variant(title, header, rows, notes_lines):
    csv_path = os.path.join(OUT, f'{title}.csv')
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)

    xlsx_path = os.path.join(OUT, f'{title}.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = '成绩'
    ws.append(header)
    for r in rows:
        ws.append(r)
    style_header(ws, len(header))
    set_widths(ws, [14] + [9] * len(SUBJECTS))
    notes = wb.create_sheet('填写说明')
    for line in notes_lines:
        notes.append([line])
    set_widths(notes, [72])
    wb.save(xlsx_path)
    print(f' - {title}.csv {os.path.getsize(csv_path)}B / {title}.xlsx {os.path.getsize(xlsx_path)}B ({len(rows)} 人)')


# 1) 姓名版（示例花名册 48 人）
write_variant(
    '成绩-测试数据',
    ['姓名'] + SUBJECTS,
    build_rows(names, blank=lambda n, s: n == '陈雨欣' and s == '物理'),
    [
        '成绩测试数据说明：',
        '1. 第一列是学生姓名（与示例花名册一致），其余列为科目成绩。',
        '2. 在应用中进入「导入数据 → 成绩」选择本文件，将自动新建一场考试。',
        '3. 「陈雨欣-物理」留空，用于测试缺考单元格的处理。',
        '4. 若导入时提示“未匹配到任何学生成绩”，说明当前花名册姓名已改动，请改用「学籍号版」。'
    ]
)

# 2) 学籍号版（首列学籍号 20260001 起，与示例花名册顺序一致）
write_variant(
    '成绩-测试数据-学籍号版',
    ['学籍号'] + SUBJECTS,
    build_rows([f'2026{str(i + 1).zfill(4)}' for i in range(len(names))],
               blank=lambda n, s: n == '20260038' and s == '物理'),
    [
        '成绩测试数据说明（学籍号版）：',
        '1. 首列是学籍号（20260001 起，与示例花名册顺序一致），其余列为科目成绩。',
        '2. 即使改过学生姓名，只要学籍号没变也能匹配。',
        '3. 「20260038（陈雨欣）-物理」留空，用于测试缺考单元格的处理。'
    ]
)

# 3) 花名册测试版（与 花名册-测试数据.csv 的 26 人配套）
if roster_names:
    write_variant(
        '成绩-测试数据-花名册测试版',
        ['姓名'] + SUBJECTS,
        build_rows(roster_names, blank=lambda n, s: n == '谢文博' and s == '历史'),
        [
            '成绩测试数据说明（花名册测试版）：',
            '1. 与「花名册-测试数据.csv」的 26 名学生配套，姓名完全一致。',
            '2. 如果你用花名册测试数据初始化了班级，请用本文件导入成绩。',
            '3. 「谢文博-历史」留空，用于测试缺考单元格的处理。'
        ]
    )
